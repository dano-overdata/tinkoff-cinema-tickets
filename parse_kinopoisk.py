
from bs4 import BeautifulSoup
import openpyxl
import time
from kinopoisk.movie import Movie
import undetected_chromedriver as uc


def parse_kinopoisk():
    wb = openpyxl.load_workbook('parsed.xlsx')
    ws = wb["Лист1"]
    cnt = 0
    global driver
    options = uc.ChromeOptions()
    options.add_argument(f'--no-sandbox')
    options.add_argument('--no-first-run --no-service-autorun --password-store=basic')
    options.add_argument("--start-maximized")
    options.add_argument('--headless')
    options.add_argument(f'--disable-gpu')
    options.add_argument(f'--disable-dev-shm-usage')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.0 Safari/605.1.15'
    options.add_argument('User-Agent={0}'.format(user_agent))
    driver = uc.Chrome()
    for film in ws['A': 'A']:
        cnt += 1
        urlNUMB = f'https://www.google.ru/search?q={str(film.value)}+%D0%BA%D0%B8%D0%BD%D0%BE%D0%BF%D0%BE%D0%B8%D1%81%D0%BA&newwindow=1&ei=MuORY8QB6-GuBNDEtZgN&ved=0ahUKEwjE4uaHjer7AhXrsIsKHVBiDdMQ4dUDCA8&uact=5&oq=1%2B1+%D0%BA%D0%B8%D0%BD%D0%BE%D0%BF%D0%BE%D0%B8%D1%81%D0%BA&gs_lcp=Cgxnd3Mtd2l6LXNlcnAQAzIFCAAQgAQyBggAEBYQHjIGCAAQFhAeOgoIABBHENYEELADOgcIABCwAxBDOgwILhDIAxCwAxBDGAE6DwguENQCEMgDELADEEMYAToHCC4Q1AIQQzoLCAAQgAQQsQMQgwE6CggAELEDEIMBEEM6BQguEIAEOgQIABBDOgoILhCxAxCDARBDOggIABCxAxCDAToICAAQgAQQsQM6CwguEIAEEMcBEK8BOhEILhCABBCxAxCDARDHARDRAzoICC4QsQMQgwE6CwguEIMBELEDEIAEOgUIABCxAzoICC4QgAQQsQM6CAgAEBYQHhAPSgQIQRgASgUIQBIBMUoECEYYAFCuBVizMmCSNGgBcAF4AIABmwKIAYITkgEGMC4xMy4zmAEAoAEByAENwAEB2gEECAEYCA&sclient=gws-wiz-serp'
        try:
            driver.get(urlNUMB)
            data = driver.page_source.encode("UTF-8").decode("UTF-8")
            soup1 = BeautifulSoup(data, "html.parser")
            page = soup1.find("h3", class_='LC20lb MBeuO DKV0Md')
            # print(page.text.split())
            for x in page.text.split():
                if '(' in x and ')' in x and len(x)==6:
                    print(x)
                    try:
                        ws.cell(row=cnt, column=2).value = x
                    except:
                        ws.cell(row=cnt, column=2).value = 'er'
            time.sleep(3)

            wb.save('parsed.xlsx')
        except Exception as ex:
            print(ex)

    driver.close()
    driver.quit()

if __name__ == '__main__':
    parse_kinopoisk()

